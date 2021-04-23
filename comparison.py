import re
import os
import sys
#import commands
import openpyxl
import subprocess
import re

from collections import defaultdict


target_dir = './target_dir'

typearmor_info = defaultdict(list)  #1 indicate it is identified as variadic function
typearmor_ind_info = dict() #address of indirect call as the key, value is the number of argument for the indirect call

function_addr_name_type = dict()
indirect_in_func = defaultdict(list)

func_name2addr = dict()
func_lineasm2llvm = dict()
func_name2line = dict()
func_addr2line = dict()
func_line2name = dict()
func_line2addr = dict()
func_icall_count = dict()
srcLine_count={}
icall_line2count = dict()

llvm_line2count = defaultdict(list)


llvm_fun_info = defaultdict(list)
llvm_ind_info = dict()

destructor_func = set()

func_addr2name = dict()

function_info = {}
source2addr = {}
ind_addr2source={}

readrdx = list() #functions which read %RDX
read63 = defaultdict(list)  #functions which read 64 bit to 32 bit
reg_bitmap=defaultdict(list)

icall_in_wrapper_taken = list() #Indirect caller in wrapper function and there is no direct caller for this wrapper function
icall_in_wrapper_dcall = list() #Indirect caller in wrapper function and there are direct callers for this wrapper function
icall_bet = list()

icall_argPointer = defaultdict(list) #indirect call whose argument is pointer to .data section
icall_argImm = defaultdict(list) #indirect call whose argument is immediate value

icall_R13 = defaultdict(list) #indirect call whose arguments read 16-bit to 32-bit register
icall_R36 = defaultdict(list) #indirect call whose arguments read 32-bit to 64-bit register

icall_xor = defaultdict(list) #Indirect callers whose arguments are passed using xor instructions


ParaWidthOB = defaultdict(list) #the recovered argument width at callee sites

ArgWidthOB = defaultdict(list)  #the recovered argument width at caller sites

ParaWidthGT = defaultdict(list) #the groud truth of the argument width at the callee site

ArgWidthGT = defaultdict(list) #the groud truth of the argument width at the caller site

ArgWidthGT_line = defaultdict(list)

variadic_over = set() #variadic functions whose number of arguments are overestimated




def find_last_colon(str_name):
	colon_index=0
	for index in range(len(str_name)-1,0,-1):
		if str_name[index] == ':':
			
			return index
	#return 0
	return colon_index

def find_dot(str_name):
	for index in range(len(str_name)):
		if str_name[index] == '.':
			return index
	return 0


#obtain the function name, start and end address by reading the symbol table
def get_function_info(binname):
	cmd = "readelf -s " +binname +" |grep FUNC|awk '{print $2,$3, $8}'"
	with os.popen(cmd) as file:
		for line in file:
			line = line.split()
			func_name = line[2]
			#print func_name
			func_start = int(line[0],16)
			func_end = func_start + int(line[1],16)
			#print func_start,func_end
			if (func_start > 0 and func_start != func_end):
				function_info[func_name] = (func_start,func_end)


def readAsm(asmfile,file_name):
	func_name_pattern = re.compile("[0-9a-fA-F]{1,16}\s<\S*>:")
	#source_name_pattern = re.compile("/\S*")
	linecount={}
	count = 0
	func_start = 0
	func_end = 0
	source_line = ''
	func_line = ''
	func_name_iner = " "
	find_bracket = 0

	source_inner = " "
	with open(asmfile,"r") as my_file:
		
		for line in my_file:
			m = func_name_pattern.match(line)

			line = line.rstrip()

			source_file_asm = line


			if 'gcc/O' in source_file_asm:
				source_file_asm_list = source_file_asm.split('gcc')

				line = source_file_asm_list[0]+'clang/O0'+source_file_asm_list[1][3:]

			if ( m != None ):
				#print "find function",line
				first = 0
				find_bracket = 0
				linecount={}
				line_m = line.split(" ")
				func_addr = int(line_m[0],16)
				func_name = line_m[1][1:len(line_m[1])-2]

				
				if "constprop" in func_name:
					dot_index = func_name.find(".constprop")
					func_name = func_name[:dot_index]
					#print "constprop **",func_name
				
				func_addr2name[func_addr] = func_name
				if func_name in function_info:
					func_start = function_info[func_name][0]
					func_end = function_info[func_name][1]


			
			
			if "/" in line and first == 0:
				func_line = line.rstrip() +";"+func_name
				

				func_addr2line[func_addr] = func_line
				func_line2name[func_line] = func_name
				if func_line not in func_line2addr:
					func_line2addr[func_line] = func_addr
				first = 1
			
				

			if "callq  *" in line:
				#print line
				line_q = line.split(":")
				addr = int(line_q[0],16)
				#print(hex(addr))

				cmd = "addr2line -e " + file_name + " " + str(hex(addr))

				result = subprocess.check_output(cmd, shell=True)

				#result2 = commands.getoutput(cmd).rstrip()

				result = result.rstrip()

				result2 = result.split()[0].decode()
				#print(result2)

				info = result2.split(":")

				line_num = info[len(info) - 1]

				if line_num.isdigit():
					source_line = result2

				

				source_file_asm = source_line

				if 'gcc/O' in source_file_asm:
					source_file_asm_list = source_file_asm.split('gcc')

					source_line = source_file_asm_list[0]+'clang/O0'+source_file_asm_list[1][3:]
				

				if addr not in ind_addr2source:
					#ind_addr2source[addr] = (source_inner,func_line,source_line_new)
					ind_addr2source[addr] = (func_name, source_line)
					 


				

			
def parseAsm(file_name):
	
	asmfile = file_name+"-dis.txt"
	cmd = "objdump -ld " + file_name + " > " +asmfile
	os.system(cmd)
	get_function_info(file_name)
	readAsm(asmfile,file_name)

#parse info generated by typearmor
def read_typearmor_file(file_name):
	#global typearmor_info, typearmor_ind_info
	with open(file_name,"r") as f:
		lines = f.readlines() 
		line_count = 0
		#for line_count in range(len(lines)):
		line = lines[line_count]

		while ("[varargs]" not in line):
			line_count += 1
			line = lines[line_count]
		
		if "[varargs]" in lines[line_count]:
			line_count += 1
			next_line = lines[line_count]
	
			while ("[args]" not in next_line):
				splitted_next_line = next_line.split()
				if len(splitted_next_line) > 1:
					function_addr = splitted_next_line[0]
					function_name = splitted_next_line[3]
					parameter_number = splitted_next_line[2]
					
					real_function_name = function_name[1:len(function_name)-1]

					if "variadic_over" in next_line:
						variadic_over.add(int(function_addr,16))

					#if real_function_name not in typearmor_info:
					typearmor_info[int(function_addr,16)].append(int(parameter_number))
					typearmor_info[int(function_addr,16)].append(1)
					#function_name_addr[real_function_name] = int(function_addr,16)
					function_addr_name_type[int(function_addr,16)] = real_function_name
					#print function_name, parameter_number
				line_count += 1
				next_line = lines[line_count]
		if "[args]" in lines[line_count]:
			#line_count += 1
			next_line = lines[line_count]
			while ("[icall-args]" not in next_line):
				splitted_next_line = next_line.split()
				#print next_line
				if len(splitted_next_line) > 1:
					function_addr = int(splitted_next_line[0],16)
					#function_name = splitted_next_line[3]
					parameter_number = int(splitted_next_line[2])

					function_name_start_index = next_line.find('(')
					function_name_end_index = next_line.rfind(')')

					real_function_name = next_line[function_name_start_index+1:function_name_end_index]


					width_info = next_line[function_name_end_index+1:].split()
					#print width_info, next_line

					if len(width_info) > (parameter_number):
						for i in range(parameter_number):
							parameterWidth = int(width_info[i])
							ParaWidthOB[function_addr].append(parameterWidth)

					if len(width_info) > (parameter_number):
						for i in range(parameter_number,len(width_info)):
							if "readRDX" in width_info[i]:
								readrdx.append(function_addr)
							if "read63" in width_info[i]:
								r13s= width_info[i].split('{')
								r13 = r13s[1]
								r13_index = r13.split(',')
								for j in range(len(r13_index)-1):
									index = int(r13_index[j])
									if index not in read63[function_addr]:
									#icall_R13[indirect_addr].append(index)
										read63[function_addr].append(index)
							if "reg_bitmap:" in width_info[i]:
								bitmap = width_info[i].split(':')[1]
								print(bitmap)
								for j in range(len(bitmap)):
									reg_bitmap[function_addr].append(int(bitmap[j]))
							

					if function_addr not in function_addr_name_type:
						function_addr_name_type[function_addr] = real_function_name
					if function_addr not in typearmor_info:
						typearmor_info[function_addr].append(parameter_number)
						typearmor_info[function_addr].append(0)
						#print real_function_name,parameter_number
				line_count += 1
				if line_count < len(lines):
					next_line = lines[line_count]
				else:
					line_count = line_count - 1
					break
		if "[icall-args]" in lines[line_count]:
			#line_count += 1
			next_line = lines[line_count]
			while ("[plts]" not in next_line ):
				splitted_next_line = next_line.split()
				if len(splitted_next_line) > 1:
					indirect_addr = int(splitted_next_line[0],16)
					function_name = splitted_next_line[3][::-1]
					argument_number = int(splitted_next_line[2])
					index = find_dot(function_name)
					function_name = function_name[index+1:]
					real_function_name = function_name[::-1][1:len(function_name)]


					function_name_start_index = next_line.find('(')
					function_name_end_index = next_line.rfind(')')

					width_info = next_line[function_name_end_index+1:].split()
					#print width_info,next_line
					if len(width_info) > (argument_number):
						for i in range(argument_number):
							if len(width_info) >= argument_number:
								if width_info[i].isdigit():
									argumentWidth = int(width_info[i])
									ArgWidthOB[indirect_addr].append(argumentWidth)
					
					if len(width_info) > (argument_number):
						for i in range(argument_number,len(width_info)):
							if "Icall" in width_info[i]:
								icall_in_wrapper_taken.append(indirect_addr)
							if "Dcall" in width_info[i]:
								icall_in_wrapper_dcall.append(indirect_addr)
							if "bet" in width_info[i]:
								icall_bet.append(indirect_addr)
							
							if "ImmArg" in width_info[i]:
								immargs= width_info[i].split('{')
								immarg = immargs[1]
								immarg_index = immarg.split(',')
								for j in range(len(immarg_index)-1):
									index = int(immarg_index[j])
									if index not in icall_argImm[indirect_addr]:
										icall_argImm[indirect_addr].append(index)
							
							if "PointerArg" in width_info[i]:
								pointerargs= width_info[i].split('{')
								pointerarg = pointerargs[1]
								pointerarg_index = pointerarg.split(',')
								for j in range(len(pointerarg_index)-1):
									index = int(pointerarg_index[j])
									if index not in icall_argPointer[indirect_addr]:
										icall_argPointer[indirect_addr].append(index)
							
							if "13Arg" in width_info[i]:
								args13= width_info[i].split('{')
								arg13 = args13[1]
								arg13_index = arg13.split(',')
								for j in range(len(arg13_index)-1):
									index = int(arg13_index[j])
									if index not in icall_R13[indirect_addr]:
										icall_R13[indirect_addr].append(index)
							
							if "36Arg" in width_info[i]:
								args36= width_info[i].split('{')
								arg36 = args36[1]
								arg36_index = arg36.split(',')
								for j in range(len(arg36_index)-1):
									index = int(arg36_index[j])
									if index not in icall_R36[indirect_addr]:
										icall_R36[indirect_addr].append(index)

							if "Xor" in width_info[i]:
								argsxor= width_info[i].split('{')
								argxor = argsxor[1]
								argxor_index = argxor.split(',')
								for j in range(len(argxor_index)-1):
									index = int(argxor_index[j])
									if index not in icall_xor[indirect_addr]:
										icall_xor[indirect_addr].append(index)
							
 
					if indirect_addr not in indirect_in_func[real_function_name]:
						indirect_in_func[real_function_name].append(indirect_addr)
					
					
					typearmor_ind_info[indirect_addr] = argument_number
				line_count += 1
				if line_count < len(lines):
					next_line = lines[line_count]
				else:
					line_count = line_count - 1
					break



#parse info generated by llvm
def read_llvm_file(file_name):
	func_llvm2line = dict()
	with open(file_name,"r") as f:
		lines = f.readlines() 
		#line_count = 0
		object_file_info=defaultdict(list)
		for line_count in range(len(lines)):
			line = lines[line_count]
			line_count +=1

			if "Function:" in line:
				line = line.rstrip('\n')
				line = line.split()          
				if len(line) > 1:
					function_name = line[1]

					'''
					old_function_name = function_name
					if function_name.find('.') != -1:
						
						number = function_name[function_name.find('.')+1:]
						if number.isdigit():
							function_name = function_name[:function_name.find('.')]
					'''

					source_file = line[len(line)-1]

					if source_file.isdigit():
						source_file = ""
						source_info = source_file
						source_file_line = 0
					else:
						source_file_index = find_last_colon(source_file)
						if source_file_index != -1:
							source_info = source_file[:source_file_index]
							source_file_line = int(source_file[source_file_index+1:])
						else:
							source_info = source_file
							source_file_line = 0

						
					parameter_number = line[len(line)-2]
					
					found = 0

					for line_asm in func_line2name:
						asm_name = func_line2name[line_asm]
						asm_comma = line_asm.find(";")
						
						
						if  asm_comma != -1:
							asm_file_index = find_last_colon(line_asm[:asm_comma])
							if asm_file_index != -1:
								asm_info = line_asm[:asm_file_index]
								
								asm_file_line = line_asm[asm_file_index+1:asm_comma]
								if asm_file_line.isdigit():
									asm_file_line = int(asm_file_line)
								else:
									asm_file_line = 0
							else:
								asm_info = line_asm
								asm_file_line
						else:
							asm_info = line_asm
							asm_file_line = 0

						c = source_file_line - asm_file_line
						#print asm_info,source_info


						index = asm_info.find(source_info)

						if source_info == asm_info:
							if abs(c) < 10 and asm_name == function_name:
								if line_asm in func_line2addr:
									func_addr = func_line2addr[line_asm]

									parameter_number = int(parameter_number)
									for j in range(len(line)-2):
										width_str = line[2+j]
										if width_str in ["*", "long"]:
											width = 8
											ParaWidthGT[func_addr].append(width)
										elif width_str == "int":
											width = 4
											ParaWidthGT[func_addr].append(width)
										elif width_str == "short":
											width = 2
											ParaWidthGT[func_addr].append(width)
										elif width_str in ["char","_Bool"]:
											width = 1
											ParaWidthGT[func_addr].append(width)
										else:
											j += 1
										

									llvm_fun_info[func_addr].append(parameter_number)


									if "variadic" in line[len(line)-3]:
										llvm_fun_info[func_addr].append(1)
									else:
										llvm_fun_info[func_addr].append(0)
									found = 1
									break 
							 

					if line_count < len(lines):
						next_line = lines[line_count]
					   
						#srcLine_info={}
						icall_count = 0
						while ("Ind-call: " in next_line):
							indirect_info = next_line.rstrip("\n")
							indirect_info = next_line.split()
							icall_count += 1

							ind_arg_count = int(indirect_info[len(indirect_info)-2])
							
							srcLine = indirect_info[len(indirect_info)-1].rstrip()

							if (function_name, srcLine) not in llvm_ind_info:
								llvm_ind_info[(function_name, srcLine)] = ind_arg_count

							
								for j in range(len(indirect_info)-2):
									width_str = indirect_info[1+j]
									
									if width_str in ["*", "long"]:
										width = 8
										ArgWidthGT[(function_name, srcLine)].append(width)
									elif width_str == "int":
										width = 4
										ArgWidthGT[(function_name, srcLine)].append(width)
									elif width_str == "short":
										width = 2
										ArgWidthGT[(function_name, srcLine)].append(width)
									elif width_str in ["char","_Bool"]:
										width = 1
										ArgWidthGT[(function_name, srcLine)].append(width)
									else:
										j += 1
									

							else:
								icall_arg = llvm_ind_info[(function_name, srcLine)]

								if icall_arg != ind_arg_count:
									llvm_ind_info.pop((function_name, srcLine))
									if (function_name, srcLine) in ArgWidthGT:
										ArgWidthGT.pop((function_name, srcLine),None)

							line_count +=1 
							if line_count < len(lines):
								next_line = lines[line_count]

						

def compare_function():
	variadic_as_normal = 0
	variadic_as_normal_fun=list()
	normal_as_variadic = 0
	normal_as_variadic_fun = list()
	underestimation = list()
	overestimation = list()
	correct=list()
	underestimation_zero_parameter = list()
	underestimation_one = list()
	width_underestimization = defaultdict(list)
	width_overestimization = defaultdict(list)
	width_collect = defaultdict(list)
	base_count = 0

	
	for i in llvm_fun_info:
		#print i
		if i in typearmor_info:
			
			llvm_count = llvm_fun_info[i][0]
			llvm_variadic = llvm_fun_info[i][1]
			typearmor_count = typearmor_info[i][0]
			typearmor_variadic = typearmor_info[i][1]

			if llvm_variadic == 0 and typearmor_variadic == 1:
				#if i == 'htab_create_alloc':
					#print '****'
				normal_as_variadic_fun.append(i)
				normal_as_variadic += 1
			if llvm_variadic == 1 and typearmor_variadic == 0:
				variadic_as_normal_fun.append(i)
				variadic_as_normal += 1
			
			if llvm_count > typearmor_count:
				base_count = typearmor_count
				if llvm_count <= 6:
					underestimation.append(i)
				elif typearmor_count < 6:
					underestimation.append(i)
			
			elif llvm_count < typearmor_count:
				base_count = llvm_count
				overestimation.append(i)
			else:
				correct.append(i)
				base_count = llvm_count
			if typearmor_count == 0 and llvm_count > typearmor_count:
				if typearmor_variadic == 0:
					underestimation_zero_parameter.append(i)
			if typearmor_count == 1 and llvm_count > typearmor_count:
				underestimation_one.append(i)
			
		   
			#compare width
			if i in ParaWidthGT:
				if i in ParaWidthOB:
					for q in range(base_count):
						widthGT = ParaWidthGT[i][q]
						widthOB = ParaWidthOB[i][q]
						if widthOB == 9:
							widthOB = 0

						if widthGT > widthOB:
							width_underestimization[i].append(q+1)
						
						if widthGT < widthOB:
							width_overestimization[i].append(q+1)

						if widthGT == widthOB:
							if i not in width_collect:
								width_collect[i].append(q+1)



		  
	return width_collect, correct,variadic_as_normal_fun, normal_as_variadic_fun, underestimation, overestimation,underestimation_zero_parameter, underestimation_one,width_overestimization,width_underestimization



def compare_ind(out, bin_name):
	underestimation = {}
	overestimation = {}
	icall_num_correct = defaultdict(list)
	icall_width_correct = defaultdict(list)
	no_ground_truth = list() 
	width_underestimization = defaultdict(list)
	width_overestimization = defaultdict(list)
	#done = ditc()
	
	for addr in ind_addr2source:
		
		(func_name, srcLine) = ind_addr2source[addr]


		srcLine_old  = srcLine
		find = 0
		base_count = 0

		if addr in typearmor_ind_info:
			typearmor_count = typearmor_ind_info[addr]

			if (func_name, srcLine) in  llvm_ind_info:
				
				llvm_count = llvm_ind_info[(func_name, srcLine)]
				

				
				if llvm_count > typearmor_count and typearmor_count<6:
					#base_count = typearmor_count
					underestimation[addr]=(llvm_count,typearmor_count)
				elif llvm_count < typearmor_count:
					#base_count = llvm_count
					overestimation[addr] = (llvm_count,typearmor_count)
				if llvm_count == typearmor_count:
					icall_num_correct[addr].append(llvm_count)
				#else:
					#base_count = typearmor_count

				base_count = min(llvm_count,typearmor_count)
				#compare argument width 

				if addr in ArgWidthOB:
					if (func_name, srcLine) in ArgWidthGT:
						for i in range(base_count):
							#print "***",src_t,ArgWidthGT[src_t][i]
							if ArgWidthGT[(func_name, srcLine)][i] < ArgWidthOB[addr][i]:
								width_overestimization[addr].append(i+1)
							elif ArgWidthGT[(func_name, srcLine)][i] > ArgWidthOB[addr][i]:
								width_underestimization[addr].append(i+1)
							else:
								icall_width_correct[addr].append(i+1)

				
				

	return icall_width_correct, icall_num_correct, underestimation,overestimation,no_ground_truth,width_overestimization,width_underestimization


# main function
if __name__ == '__main__':

	#execute groundTruth.py to obtain the ground truth first

	if len(sys.argv) < 4:
		print("please set the binary path, groundTruth folder, book_name path and sheet_name")
		exit(1)

	bin_path = sys.argv[1]
	#target_dir = sys.argv[2]
	groundTruth = sys.argv[2]

	book_name = sys.argv[3]

	sheet_name = sys.argv[4]

	
	
	
	
	pattern = re.compile(r"\s*\[\s*(?P<nr>[\d]{1,2})\]\s*"
					"(?P<name>[\S]+)\s*"
					"(?P<type>[\S]+)\s*"
					"(?P<address>[\S]+)\s*"
					"(?P<offset>[\S]+)\s*"
					"(?P<size>[\S]+)\s*"
					"[^\n]*$")
	
	#get text size of the binary
	cmd = "readelf -S " +bin_path
	with os.popen(cmd) as file:
		lines = file.readlines()
		for i in range(len(lines)-1):
			if i %2 == 1:
				line = lines[i].strip()
				next_line = lines[i+1].strip()
				total_line = line +" " + next_line
				m = pattern.match(total_line)
				if((m != None) and (m.group('name') == ".text")):
					text_size = int(m.group("size"),16)
	
	

	

	orig_bin = os.path.basename(bin_path)



	llvm_path = orig_bin + "-llvm-info.txt"
	typearmor_path = orig_bin + "-typearmor.txt"

	variadic_as_normal = "variadic.txt"
	out_var = open(variadic_as_normal,"a")

	
	cmd = "mkdir -p " + target_dir
	os.system(cmd)

	
	cmd = "mkdir -p " + target_dir + "/" + orig_bin
	os.system(cmd)
	
	
	cwd = os.getcwd()
	type_armor_path = cwd+"/typearmor-master"
	os.chdir(type_armor_path+"/server-bins")

	print(type_armor_path)

	
	
	cmd = "cp " + bin_path + " " + orig_bin
	os.system(cmd)


	cmd = "rm -f "+orig_bin+".o"
	os.system(cmd)
	
	
	 
	cmd = "bash "+type_armor_path+"/run-ta-static.sh " + orig_bin
	os.system(cmd)
	
	
	
	os.chdir(cwd+"/"+target_dir + "/" + orig_bin)

	cmd = "mv " +type_armor_path+"/out/" + "binfo."+orig_bin +" "  + typearmor_path
	os.system(cmd)

	#cmd = "cp " + orig_bin + " "  + target_dir + "/" + orig_bin +"/" + orig_bin
	#os.system(cmd)

	
	cmd = "cp " + bin_path + " " + orig_bin
	os.system(cmd)
	

	if orig_bin.find("clang") != -1:
	
		cmd = "cp " + groundTruth+"/"+ llvm_path +" "  +llvm_path
		os.system(cmd)

	elif orig_bin.find("gcc") != -1:
		index = orig_bin.find("-gcc-O")

		llvmName = orig_bin[:index]+"-clang-O0-llvm-info.txt"

		cmd = "cp " + groundTruth+"/" + llvmName + " "+ llvm_path
		os.system(cmd)


	parseAsm(orig_bin)
	
	read_typearmor_file(typearmor_path)
	read_llvm_file(llvm_path)


	
	outfile = "ArgCount.txt"
	out = open(outfile,"w")

	outfile_width = "ArgWidth.txt"
	out_width = open(outfile_width,"w")

	
	correct_func_width, correct, variadic_as_normal_fun, normal_as_variadic_fun, fun_underestimation, fun_overestimation,fun_underestimation_zero,underestimation_one, width_overestimization, width_underestimization = compare_function()
	correct_icall_width, icall_num_correct,  underestimation_ind,overestimation_ind,no_ground_truth,ind_width_overestimization, ind_width_underestimization = compare_ind(out,orig_bin)


	width_correct = 0
	icall_width_correct = 0



	for i in correct_func_width:
		if i in correct:
			if len(correct_func_width[i]) == llvm_fun_info[i][0]:
				width_correct += 1

	for i in correct_icall_width:
		if i in icall_num_correct:
			if len(correct_icall_width[i]) == icall_num_correct[i]:
				icall_width_correct += 1




	count_normal_to_variadic = 0
	count_varaidc_to_normal = 0
	count_varadic_paramOver = 0
	count_readRdx = 0
	count_temp = 0
	count_icall_dcall = 0
	count_icall_taken = 0
	count_not_used = 0
	count_icall_imm = 0
	count_icall_pointer = 0
	count_lea = 0
	count_null = 0
	count_prom = 0
	count_paraWidth_underestimization_not_use = 0
	count_paraWidth_underestimization_other = 0
	count_paranum_under = 0
	count_paraWidth_overestimization_other = 0
	count_argunder = 0
	count_paramover = 0
	count_destructor = 0
	count_icall_bet = 0
	count_imm_over = 0
	icall_imm_over = 0
	icall_pointer_over = 0
	icall_xor_over = 0
	callee_arg_width_over_48 = 0
	callee_arg_width_prop = 0
	callee_arg_width_under = 0

	push_under = 0
	rdx_under = 0

	count_rdx_total = len(readrdx)
	count_push_total = len(read63)


	if len(variadic_as_normal_fun) > 0:
		out.write("variadic as normal function\n")
		for i in variadic_as_normal_fun:
			count_varaidc_to_normal += 1
			out.write(str(hex(i))+"\n")
			out_var.write(str(hex(i))+"\n")
	out.write("\n")
	if len(normal_as_variadic_fun) > 0:
		out.write("normal function as variadic function\n")
		for i in normal_as_variadic_fun:
			count_normal_to_variadic += 1
			out.write(str(hex(i))+" ")

			if i in func_addr2line:
				out.write(func_addr2line[i])
			
			out.write("\n")
			out_var.write(str(hex(i))+"\n")
	out.write("\n")
	if len(fun_overestimation) > 0:
		out.write("function parameter overestimation, serious\n")
		for i in fun_overestimation:
			count_paramover += 1
			if i in func_addr2name:
				name = func_addr2name[i]
				if name in destructor_func:
					count_destructor += 1
			out.write(str(hex(i))+" ")
			if i in variadic_over:
				count_varadic_paramOver += 1
				out.write("variadic_over ")
			elif i in readrdx:
				out.write("readrdx ")
				count_readRdx += 1
			else:
				out.write("("+str(typearmor_info[i][0])+" "+str(llvm_fun_info[i][0])+")")
			if i in func_addr2line:
				out.write(func_addr2line[i])
			out.write("\n")
	out.write("\n")
	

	if len(correct) > 0:
		count = 0
		out.write("function parameter correct\n")
		for i in correct:
			out.write(str(hex(i))+" ")
			out.write("("+str(typearmor_info[i][0])+" "+str(llvm_fun_info[i][0])+")")
			out.write("write(")
			for index in range(typearmor_info[i][0],6):
				#print(index)
				if index < 6 and i in reg_bitmap:
					if reg_bitmap[i][index] != 0:
						out.write(str(index+1)+" ")
			out.write(")")
				

			if i in func_addr2line:
				out.write(func_addr2line[i])
			
			out.write("\n")
		
		
	out.write("\n")



	if len(fun_underestimation) > 0:
		count = 0
		out.write("function parameter underestimation\n")
		for i in fun_underestimation:
			if i not in normal_as_variadic_fun:
				count_not_used += 1
				out.write(str(hex(i))+" ")
				out.write("("+str(typearmor_info[i][0])+" "+str(llvm_fun_info[i][0])+")")
				out.write("write(")
				for index in range(typearmor_info[i][0],llvm_fun_info[i][0]):
					#print(index)
					if index < 6 and i in reg_bitmap:
						if reg_bitmap[i][index] != 0:
							out.write(str(index+1)+" ")
				out.write(")")

				if i in readrdx:
					rdx_under += 1

				if i in func_addr2line:
					out.write(func_addr2line[i])
				
				out.write("\n")
			if i in normal_as_variadic_fun:
				count += 1
		out.write("underestimation caused by treating noraml to variadic " + str(count) +"\n")
	out.write("\n")
	

	if len(underestimation_ind) > 0:
		out.write("indirect call argument underestimation, serious\n")
		for i in underestimation_ind:
			count_argunder += 1
			out.write(str(hex(i))+" ")
			out.write(str(underestimation_ind[i][0])+" "+str(underestimation_ind[i][1]))
			if i in ind_addr2source:
				out.write(ind_addr2source[i][1])
			out.write("\n")
	out.write("\n")
	count_num_over = 0
	if len(overestimation_ind) > 0:
		out.write("indirect call argument overestimation\n")
		for i in overestimation_ind:
			count_num_over += 1
			out.write(str(hex(i))+" ")
			out.write(str(overestimation_ind[i][0])+" "+str(overestimation_ind[i][1]))
			if i in icall_in_wrapper_dcall:
				out.write(" wrapper_dcall ")
				count_icall_dcall += 1
			elif i in icall_in_wrapper_taken:
				out.write(" wrapper_icall ")
				count_icall_taken += 1
			else:
				count_temp += 1
				if i in icall_argImm:
					count_imm_over += 1
			if i in icall_bet:
				count_icall_bet += 1

			if i in ind_addr2source:
				out.write(ind_addr2source[i][1])

			out.write("\n")
	out.write("\n")
	if len(no_ground_truth) > 0:
		out.write("indirect call does not have ground truth\n")
		for i in no_ground_truth:
			out.write(str(hex(i))+"\n")
	
	

	if len(width_overestimization) > 0:
		out_width.write("Function parameter width overestimization, serious\n")
		for i in width_overestimization:
			find = 0
			out_width.write(str(hex(i))+" ")
			for j in width_overestimization[i]:
				out_width.write(str(j) + " ")
				out_width.write("("+str(ParaWidthGT[i][j-1])+" "+str(ParaWidthOB[i][j-1])+")")
				if ParaWidthGT[i][j-1] == 4:
					find = 1
				else:
					find =2
			if find == 1 and i not in read63:
				callee_arg_width_over_48 = callee_arg_width_over_48 + 1
			if find == 2:
				callee_arg_width_prop = callee_arg_width_prop + 1

			
			#out_width.write("\n")
			if i in read63:
				
				out_width.write("read63(")
				s = " "
				for j in read63[i]:
					s += str(j)
					out_width.write(str(j)+" ")
				out_width.write(")")
				if s!=" ":
					count_lea += 1
			else:
				count_paraWidth_overestimization_other += 1
			
			out_width.write("\n")

		out_width.write("\n")
	
	if len(width_underestimization) > 0:
		out_width.write("Function parameter width underestimization\n")
		for i in width_underestimization:
			find = 0
			find_push_under = 0
			if i in fun_underestimation:
				count_paraWidth_underestimization_not_use += 1
			else:
				count_paraWidth_underestimization_other += 1
			if i in read63:
				for j in read63[i]:
					if ParaWidthGT[i][j-1] > ParaWidthOB[i][j-1]:
						find_push_under = 1
			if find_push_under:
				push_under += 1
			out_width.write(str(hex(i))+" ")
			for j in width_underestimization[i]:
				out_width.write(str(j) + " ")
				if ParaWidthOB[i][j-1] != 9:
					find = 1
					out_width.write("("+str(ParaWidthGT[i][j-1])+" "+str(ParaWidthOB[i][j-1])+")")
				else:
					out_width.write("(null)")
			if find == 1:
				callee_arg_width_under = callee_arg_width_under + 1
			out_width.write("\n")
		out_width.write("\n")
	
	count_icall_width_over_other = 0

	if len(ind_width_overestimization) > 0:
		out_width.write("icall argument width overestimization\n")
		for i in ind_width_overestimization:
			out_width.write(str(hex(i))+" ")
			for j in ind_width_overestimization[i]:
				out_width.write(str(j) + " ")
			# out_width.write("\n")
			if i in icall_R13:
				
				s = " "
				out_width.write("icallR13(")
				for t in icall_R13[i]:
					if t in ind_width_overestimization[i]:
						s += str(t)
						out_width.write(str(t) + " ")
				out_width.write(")")
				if s != " ":
					count_prom += 1

			if i in icall_R36:
				s = " "
				out_width.write("icallR36(")
				for t in icall_R36[i]:
					if t in ind_width_overestimization[i]:
						s += str(t)
						out_width.write(str(t) + " ")
				out_width.write(")")
				if i not in icall_R13 and s != " ":
					count_prom += 1
			
			if i in icall_argImm:
				s = " "
				out_width.write("IcallImm(")
				for t in icall_argImm[i]:
					if i in icall_xor:
						if t in icall_xor[i]:
							continue
					if t in ind_width_overestimization[i]:
						s += str(t)
						out_width.write(str(t) + " ")
				out_width.write(")")
				if s != " ":
					icall_imm_over += 1 
			
			if i in icall_argPointer:
				s = " "
				out_width.write("IcallPointer(")
				for t in icall_argPointer[i]:
					if t in ind_width_overestimization[i]:
						s += str(t)
						out_width.write(str(t) + " ")
				out_width.write(")")
				if s != " ":
					icall_pointer_over += 1
			
			if i in icall_xor:
				s = " "
				out_width.write("Icallxor(")
				for t in icall_xor[i]:
					if t in ind_width_overestimization[i]:
						s += str(t)
						out_width.write(str(t) + " ")
				out_width.write(")")
				if s != " ":
					icall_xor_over += 1

			if i not in icall_R13 and i not in icall_R36:
				count_icall_width_over_other += 1

			if i in ind_addr2source:
				out_width.write(ind_addr2source[i][1])
		
			out_width.write("\n")
		out_width.write("\n")
	
	count_icall_width_under_other = 0
	count_icall_xor = 0
	count_both_two = 0
	count_both_three = 0
	count_width_under_toal = 0
	if len(ind_width_underestimization) > 0:
		out_width.write("icall argument width underestimization, serious\n")
		for i in ind_width_underestimization:
			count_width_under_toal += 1
			out_width.write(str(hex(i))+" ")
			for j in ind_width_underestimization[i]:
				out_width.write(str(j) + " ")
			#out_width.write("\n")

			if i in icall_argPointer:
				s = " "
				
				out_width.write("icallPointer(")
				for t in icall_argPointer[i]:
					#print t,type(t)
					if t in ind_width_underestimization[i]:
						s += str(t)
						out_width.write(str(t) + " ")
				out_width.write(") ")
				if s!= " ":
					count_icall_pointer += 1
			if i in icall_argImm:
				#count_icall_imm += 1
				s = " "
				out_width.write("icallImm(")
				for t in icall_argImm[i]:
					if i in icall_xor:
						if t in icall_xor[i]:
							continue
					if t in ind_width_underestimization[i]:
						s += str(t)
						out_width.write(str(t) + " ")
				out_width.write(") ")
				if s != " ":
					count_icall_imm += 1


			if i in icall_xor:
				
				s = " "
				out_width.write("icallXor(")
				for t in icall_xor[i]:
					if t in ind_width_underestimization[i]:
						s += str(t)
						out_width.write(str(t) + " ")
				out_width.write(")")
				if s != " ":
					count_icall_xor += 1
			'''
			if i in [icall_argPointer,icall_argImm] or i in [icall_argPointer,icall_xor] or i in [icall_argImm,icall_xor]:
				count_both_two += 1
			
			if i in [icall_argPointer, icall_argImm, icall_xor]:
				count_both_three += 1
			'''

			if i not in icall_argPointer and i not in icall_argImm and i not in icall_xor:
				count_icall_width_under_other += 1

			if i in ind_addr2source:
				out_width.write(ind_addr2source[i][1])
			
			out_width.write("\n")

	
	out.close()
	out_var.close()
	out_width.close()




	count_icall_pointer_total = len(icall_argPointer)
	count_icall_imm_total = len(icall_argImm)
	count_icall_xor_total = len(icall_xor)
	count_lea_total = len(read63)
	count_read_rdx_total = len(readrdx)
	count_icall_in_wrapper_dcall_total = len(icall_in_wrapper_dcall)
	count_icall_in_wrapper_taken_total = len(icall_in_wrapper_taken)
	
	count_icall_r13_total = len(icall_R13)
	count_icall_r36_total = len(icall_R36)

	count_icall_bet_total = len(icall_bet)
	

	
	index = orig_bin.find("-clang")
	binname = orig_bin[:index]
	#sheet_name = orig_bin[index+1:]

	exists = os.path.isfile(book_name)

	if exists:
		book = openpyxl.load_workbook(book_name)
	else:
		book = openpyxl.Workbook()

	#sheet = book.get_sheet_by_name("clang-O0-binutils")
	


	if sheet_name not in book.sheetnames:
		book.create_sheet(sheet_name)
		sheet = book.get_sheet_by_name(sheet_name) 
		str1 = ["Application","Nor2var", "Var2Nor", "VarOver", "Unread", 
				"rdx", "read63", "Ret","Uninit", "Wrapper",
				"Unmodified", "Temp", "Imm", "Pointer", "Null", "Prom"]


		for row in range(1):
			for col in range(15):
				#print str1[col]
				sheet.cell(column=col+1, row=row+1, value=str1[col])
	else:
		sheet = book.get_sheet_by_name(sheet_name) 
	

	str_name = [binname,count_normal_to_variadic,count_varaidc_to_normal, count_varadic_paramOver, count_not_used,count_readRdx,
				count_lea, 0, 0, count_icall_in_wrapper_dcall_total+count_icall_in_wrapper_taken_total,
				count_icall_bet, count_temp, icall_imm_over, icall_pointer_over, icall_xor_over,count_prom]


	for row in range(sheet.max_row,sheet.max_row+1):
		for col in range(15):
			#print str[col]
			sheet.cell(column=col+1, row=row+1, value=str_name[col])
			

	book.save(book_name)
	