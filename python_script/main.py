import os
import sys
import argparse
import openpyxl
from multiprocessing import Process

from compare_arg import *

'''

'''

opts = ['O0','O1','O2','O3']

def get_config():
	'''
	get config information from command line
	'''
	parser = argparse.ArgumentParser()

	parser.add_argument('-s', '--source_folder', dest='source_folder', help='The folder of source code to be processed.', type=str, required=True)
	
	parser.add_argument('-opt', '--opt', dest='opt_level', help='The optimization level when compiling the source code.', type=str, required=False)
	parser.add_argument('-c', '--compiler', dest='compiler', help='The compiler that will be used to compile the source code.', type=str, required=False)
	
	parser.add_argument('-g', '--groundtruth_folder', dest='groundtruth_folder', help='The root folder that saves the ground truth for each function.', type=str, required=False)

	parser.add_argument('-oc', '--only_compare', dest='only_compare', help='The flag to indicate whether we only want to compare the result grnerated by TypeArmor.', type=str, required=True,default='false')
	parser.add_argument('-gs', '--get_signature', dest='get_signature',help='The flag to indicate whether we only want to get the result grnerated by TypeArmor.',
				type=str, required=False, default='false')

	parser.add_argument('-bn', '--book_name', dest='book_name',
                        help='The .xls file name to store the statistical result.',
                        type=str, required=False, default='false')

	args = parser.parse_args()
	
	config_info = {
		'source_folder': args.source_folder,
		'opt_level': args.opt_level,
		'compiler': args.compiler,
		'groundtruth_folder': args.groundtruth_folder,
		'only_compare':args.only_compare,
		'get_signature': args.get_signature,
		'book_name': args.book_name
	}

	return config_info


def print_func_comparison_result(func_num_underestimation, func_num_correct, func_num_overestimation, func_addr2name,
								 out, compareHandler, binary,typearmor_over48=None,typearmor_lea=None,typearmor_prom=None,ghidra='false'):
	count_varaidc_to_normal = 0
	count_normal_to_variadic = 0
	count_callee_paramover = 0
	count_callee_paramunder = 0
	count_callee_paramcorrect = 0
	count_callee_paramcorrect_0 = 0
	count_varadic_paramOver = 0
	count_readRdx = 0
	count_rdx_under = 0
	name = ' '
	var2nor_over = 0
	count_variadic_under = 0
	# print out the comparison result on the number of arguments at the callee site
	if len(compareHandler.variadic_as_normal_fun) > 0:
		out.write("\tvariadic as normal function\n")
		for i in compareHandler.variadic_as_normal_fun:
			count_varaidc_to_normal += 1
			if i in func_addr2name:
				name = func_addr2name[i]

			out.write("\t\t" + str(hex(i)) + ": " + name + ' ')

			if compareHandler.typearmor_info[i][0] > compareHandler.llvm_fun_info[i][0]:
				var2nor_over += 1
				out.write("(" + str(compareHandler.typearmor_info[i][0]) + " " + str(
					compareHandler.llvm_fun_info[i][0]) + ")")
			if i in compareHandler.func_addr2line:
				out.write(compareHandler.func_addr2line[i])

			out.write("\n")

	if len(compareHandler.normal_as_variadic_fun) > 0:
		out.write("\tnormal function as variadic function\n")
		for i in compareHandler.normal_as_variadic_fun:
			count_normal_to_variadic += 1
			if i in func_addr2name:
				name = func_addr2name[i]
			out.write("\t\t" + str(hex(i)) + ": " + name + ' ')

			if i in compareHandler.func_addr2line:
				out.write(compareHandler.func_addr2line[i])

			out.write("\n")

	if len(func_num_overestimation) > 0:
		out.write("\tnumber of function parameter overestimated, serious\n")
		for i in func_num_overestimation:
			count_callee_paramover += 1
			if i in func_addr2name:
				name = func_addr2name[i]

			out.write("\t\t" + str(hex(i)) + ": " + name)
			if i in compareHandler.variadic_over:
				count_varadic_paramOver += 1
				out.write("variadic_over ")
			elif i in compareHandler.readrdx:
				out.write("readrdx ")
				count_readRdx += 1
			else:
				out.write("(" + str(compareHandler.typearmor_info[i][0]) + " " + str(
					compareHandler.llvm_fun_info[i][0]) + ")")
			if i in compareHandler.func_addr2line:
				out.write(compareHandler.func_addr2line[i])
			out.write("\n")

	if len(func_num_underestimation) > 0:
		count = 0
		out.write("\tfunction parameter underestimation\n")
		for i in func_num_underestimation:
			count_callee_paramunder += 1
			if i in compareHandler.variadic_under:
				out.write("\t\t" + str(hex(i)) + " variadic under \n")
				count_variadic_under += 1
			elif i not in compareHandler.normal_as_variadic_fun:
				# count_not_used += 1
				out.write("\t\t" + str(hex(i)) + " ")
				if i in compareHandler.readrdx:
					out.write("readrdx ")
					count_rdx_under += 1
				
				out.write("(" + str(compareHandler.typearmor_info[i][0]) + " " + str(
					compareHandler.llvm_fun_info[i][0]) + ")")

				if i in compareHandler.func_addr2line:
					out.write(compareHandler.func_addr2line[i])

				out.write("\n")

			if i in compareHandler.normal_as_variadic_fun:
				count += 1
		out.write("\t\tunderestimation caused by treating noraml to variadic " + str(count) + "\n")

	if len(func_num_correct) > 0:
		count = 0
		out.write("\tfunction parameter correct\n")
		for i in func_num_correct:
			count_callee_paramcorrect += 1
			out.write("\t\t" + str(hex(i)) + " ")
			if i in compareHandler.llvm_fun_info:
				out.write("(" + str(compareHandler.typearmor_info[i][0]) + " " + str(compareHandler.llvm_fun_info[i][0]) + ")")
			else:
				print('cannot find ' + str(hex(i)))

			if i in compareHandler.func_addr2line:
				out.write(compareHandler.func_addr2line[i])

			if  compareHandler.llvm_fun_info[i][0] == 0:
				count_callee_paramcorrect_0 += 1

			out.write("\n")

	# print out the comparison result on the width of the argument at the callee site
	callee_arg_width_over_48 = 0
	callee_arg_width_prop = 0
	count_lea = 0
	count_paraWidth_overestimization_other = 0
	callee_arg_width_over = 0
	# THE WIDTH OF ARGUMENT IS OVERESTIMATED
	if len(compareHandler.func_width_overestimization) > 0:
		out.write("\tFunction parameter width overestimation, serious\n")
		for i in compareHandler.func_width_overestimization:
			callee_arg_width_over += 1
			find = 0
			out.write("\t\t" + str(hex(i)) + " ")
			for j in compareHandler.func_width_overestimization[i]:
				out.write(str(j) + " ")
				out.write("(" + str(compareHandler.ParaWidthGT[i][j - 1]) + " " + str(
					compareHandler.ParaWidthOB[i][j - 1]) + ")")
				if compareHandler.ParaWidthGT[i][j - 1] == 4:
					find = 1
				else:
					find = 2

			if find == 1 and i not in compareHandler.read63:
				if ghidra == 'false':
					out.write("over48")
					callee_arg_width_over_48 = callee_arg_width_over_48 + 1
				elif ghidra == 'true':
					if i in typearmor_over48:
						out.write("over48")
						callee_arg_width_over_48 = callee_arg_width_over_48 + 1
			if find == 2:
				if ghidra == 'false':
					callee_arg_width_prop = callee_arg_width_prop + 1
				elif ghidra == 'true':
					#if i in typearmor_prom:
					callee_arg_width_prop = callee_arg_width_prop + 1

			if i in compareHandler.read63 and find == 1:
				if ghidra == 'false':
					out.write("read63(")
				elif ghidra == 'true':
					if i in typearmor_over48 :
						out.write("over48")
					else:
						out.write("read63(")

						#callee_arg_width_over_48 = callee_arg_width_over_48 + 1
					#if i in typearmor_prom:
						#callee_arg_width_prop = callee_arg_width_prop + 1

				s = " "
				for j in compareHandler.read63[i]:
					s += str(j)
					out.write(str(j) + " ")
				out.write(")")
				if s != " ":
					if ghidra == 'true':
						if i in typearmor_over48:
							callee_arg_width_over_48 = callee_arg_width_over_48 + 1
						elif i in typearmor_prom:
							callee_arg_width_prop = callee_arg_width_prop + 1
						else:
							count_lea += 1
					elif ghidra == 'false':
						count_lea += 1
			else:

				count_paraWidth_overestimization_other += 1

			out.write("\n")

	# THE WIDTH OF THE ARGUMENT IS UNDERESTIMATED
	count_paraWidth_underestimization_not_use = 0
	count_paraWidth_underestimization_other = 0
	count_push_under = 0
	callee_arg_width_under = 0
	count_arg_width_correct_0 = 0
	count_arg_width_correct = 0
	if len(compareHandler.func_width_underestimization) > 0:
		out.write("\tFunction parameter width underestimization\n")
		for i in compareHandler.func_width_underestimization:
			find = 0
			find_push_under = 0
			if i in func_num_underestimation:
				count_paraWidth_underestimization_not_use += 1
			else:
				count_paraWidth_underestimization_other += 1
			if i in compareHandler.read63:
				for j in compareHandler.read63[i]:
					if i in compareHandler.ParaWidthGT:
						if j <= len(compareHandler.ParaWidthGT[i]) and j <= len(compareHandler.ParaWidthOB[i]):
							if compareHandler.ParaWidthGT[i][j - 1] > compareHandler.ParaWidthOB[i][j - 1]:
								find_push_under = 1
			if find_push_under:
				count_push_under += 1
			out.write("\t\t" + str(hex(i)) + " ")
			for j in compareHandler.func_width_underestimization[i]:
				out.write(str(j) + " ")
				if compareHandler.ParaWidthOB[i][j - 1] != 9:
					find = 1
					out.write("(" + str(compareHandler.ParaWidthGT[i][j - 1]) + " " + str(
						compareHandler.ParaWidthOB[i][j - 1]) + ")")
				else:
					out.write("(null)")
			if find == 1:
				callee_arg_width_under = callee_arg_width_under + 1
			out.write("\n")

	if len(compareHandler.func_width_correct) > 0:
		out.write("\tFunction parameter width correct\n")
		for i in compareHandler.func_width_correct:
			if i in func_num_correct:
				if len(compareHandler.func_width_correct[i]) == 1:
					if compareHandler.func_width_correct[i][0] == 0:
						count_arg_width_correct_0 += 1
						out.write("\t\t" + str(hex(i)) + " \n")
					else:
						if len(compareHandler.func_width_correct[i]) == compareHandler.typearmor_info[i][0]:
							count_arg_width_correct += 1
							out.write("\t\t" + str(hex(i)) + " \n")

				else:
					if len(compareHandler.func_width_correct[i]) == compareHandler.typearmor_info[i][0]:
						count_arg_width_correct += 1
						out.write("\t\t" + str(hex(i)) + " \n")


	result = [binary, count_variadic_under, var2nor_over, count_varaidc_to_normal, count_normal_to_variadic, count_callee_paramover,
			  count_callee_paramunder,
			  count_callee_paramcorrect,count_callee_paramcorrect_0, count_varadic_paramOver, count_readRdx, count_rdx_under, count_push_under,
			  callee_arg_width_over_48, callee_arg_width_prop, count_lea,
			  count_paraWidth_overestimization_other, callee_arg_width_over, callee_arg_width_under,
			  count_arg_width_correct, count_arg_width_correct_0,
			  count_paraWidth_underestimization_not_use, count_paraWidth_underestimization_other,
			  compareHandler.total_func, compareHandler.total_func_0, compareHandler.variadic_func]

	return result


def print_icall_comparison_result(icall_num_underestimation, icall_num_correct, icall_num_overestimation,
								  out, compareHandler, binary):
	count_icall_temp = 0
	count_icall_imm = 0
	count_icall_xor = 0
	count_icall_pointer = 0
	count_icall_paramover = 0
	count_icall_taken = 0
	count_icall_dcall = 0
	count_icall_paramunder = 0
	count_icall_paramcorrect = 0
	count_icall_paramcorrect_0 = 0
	count_icall_wrapper_under = 0
	count_temp_under = 0
	count_icall_bet = 0

	if len(icall_num_underestimation) > 0:
		print(icall_num_underestimation)
		out.write("\tindirect call argument underestimation, serious\n")
		for i in icall_num_underestimation:
			count_icall_paramunder += 1
			# if i in func_addr2name:
			# name = func_addr2name[i]
			out.write("\t\t" + str(hex(i)) + " ")
			out.write(str(icall_num_underestimation[i][0]) + " " + str(icall_num_underestimation[i][1]) + " ")
			if i in compareHandler.icall_in_wrapper_dcall:
				out.write(" wrapper_dcall ")
				count_icall_wrapper_under += 1
			elif i in compareHandler.icall_in_wrapper_taken:
				out.write(" wrapper_icall ")
				count_icall_wrapper_under += 1

			elif i in compareHandler.ghidra_icall_num_overestimation:
				count_temp_under += 1
			if i in compareHandler.ind_addr2source:
				out.write(compareHandler.ind_addr2source[i][1])
			out.write("\n")

	if len(icall_num_overestimation) > 0:
		count = 0
		out.write("\tindirect call argument overestimation\n")
		for i in icall_num_overestimation:
			count_icall_paramover += 1
			out.write("\t\t" + str(hex(i)) + " ")
			out.write(str(icall_num_overestimation[i][0]) + " " + str(icall_num_overestimation[i][1]) + " ")
			if i in compareHandler.icall_in_wrapper_dcall:
				out.write(" wrapper_dcall ")
				count_icall_dcall += 1
			elif i in compareHandler.icall_in_wrapper_taken:
				out.write(" wrapper_icall ")
				count_icall_taken += 1
			else:
				count_icall_temp += 1
				if i in compareHandler.icall_argImm:
					#count_imm_over += 1
					pass
			if i in compareHandler.icall_bet:
				# count_icall_bet += 1
				pass

			if i in compareHandler.ind_addr2source:
				out.write(compareHandler.ind_addr2source[i][1])

			out.write("\n")
	if len(icall_num_correct) > 0:
		count = 0
		out.write("\tindirect call argument  correct\n")
		for i in icall_num_correct:
			count_icall_paramcorrect += 1
			out.write("\t\t" + str(hex(i)) + " ")
			out.write(str(icall_num_correct[i][0]) + " ")
			if i in compareHandler.ind_addr2source:
				out.write(compareHandler.ind_addr2source[i][1])
			out.write("\n")

			if icall_num_correct[i][0] == 0:
				count_icall_paramcorrect_0 += 1

	# The width of the argument is overestimated at the caller site
	count_icall_width_over_other = 0
	count_prom = 0
	icall_imm_width_over = 0
	icall_pointer_width_over = 0
	icall_xor_width_over = 0
	if len(compareHandler.icall_width_overestimization) > 0:
		out.write("\ticall argument width overestimization\n")
		for i in compareHandler.icall_width_overestimization:
			out.write("\t\t" + str(hex(i)) + " ")
			for j in compareHandler.icall_width_overestimization[i]:
				out.write(str(j) + " ")
			# out_width.write("\n")
			if i in compareHandler.icall_R13:

				s = " "
				out.write("icallR13(")
				for t in compareHandler.icall_R13[i]:
					if t in compareHandler.icall_width_overestimization[i]:
						s += str(t)
						out.write(str(t) + " ")
				out.write(")")
				if s != " ":
					count_prom += 1

			if i in compareHandler.icall_R36:
				s = " "
				out.write("icallR36(")
				for t in compareHandler.icall_R36[i]:
					if t in compareHandler.icall_width_overestimization[i]:
						s += str(t)
						out.write(str(t) + " ")
				out.write(")")
				if i not in compareHandler.icall_R13 and s != " ":
					count_prom += 1

			if i in compareHandler.icall_argImm:
				s = " "
				out.write("IcallImm(")
				for t in compareHandler.icall_argImm[i]:
					if i in compareHandler.icall_xor:
						if t in compareHandler.icall_xor[i]:
							continue
					if t in compareHandler.icall_width_overestimization[i]:
						s += str(t)
						out.write(str(t) + " ")
				out.write(")")
				if s != " ":
					icall_imm_width_over += 1

			if i in compareHandler.icall_argPointer:
				s = " "
				out.write("IcallPointer(")
				for t in compareHandler.icall_argPointer[i]:
					if t in compareHandler.icall_width_overestimization[i]:
						s += str(t)
						out.write(str(t) + " ")
				out.write(")")
				if s != " ":
					icall_pointer_width_over += 1

			if i in compareHandler.icall_xor:
				s = " "
				out.write("Icallxor(")
				for t in compareHandler.icall_xor[i]:
					if t in compareHandler.icall_width_overestimization[i]:
						s += str(t)
						out.write(str(t) + " ")
				out.write(")")
				if s != " ":
					icall_xor_width_over += 1

			if i not in compareHandler.icall_R13 and i not in compareHandler.icall_R36:
				count_icall_width_over_other += 1

			if i in compareHandler.ind_addr2source:
				out.write(compareHandler.ind_addr2source[i][1])

			out.write("\n")

	# The width of the argument is underestimated at the caller site
	count_icall_width_under_other = 0
	count_icall_xor = 0
	count_both_two = 0
	count_both_three = 0
	count_icall_width_under = 0
	count_icall_pointer_width_under = 0
	count_icall_imm_width_under = 0
	count_icall_xor_width_under = 0
	count_icall_width_correct = 0
	if len(compareHandler.icall_width_underestimization) > 0:
		out.write("\ticall argument width underestimization, serious\n")
		for i in compareHandler.icall_width_underestimization:
			count_icall_width_under += 1
			out.write("\t\t" + str(hex(i)) + " ")
			for j in compareHandler.icall_width_underestimization[i]:
				out.write(str(j) + " ")
			# out_width.write("\n")

			if i in compareHandler.icall_argPointer:
				s = " "

				out.write("icallPointer(")
				for t in compareHandler.icall_argPointer[i]:
					# print t,type(t)
					if t in compareHandler.icall_width_underestimization[i]:
						s += str(t)
						out.write(str(t) + " ")
				out.write(") ")
				if s != " ":
					count_icall_pointer_width_under += 1
			if i in compareHandler.icall_argImm:
				# count_icall_imm += 1
				s = " "
				out.write("icallImm(")
				for t in compareHandler.icall_argImm[i]:
					if i in compareHandler.icall_xor:
						if t in compareHandler.icall_xor[i]:
							continue
					if t in compareHandler.icall_width_underestimization[i]:
						s += str(t)
						out.write(str(t) + " ")
				out.write(") ")
				if s != " ":
					count_icall_imm_width_under += 1

			if i in compareHandler.icall_xor:

				s = " "
				out.write("icallXor(")
				for t in compareHandler.icall_xor[i]:
					if t in compareHandler.icall_width_underestimization[i]:
						s += str(t)
						out.write(str(t) + " ")
				out.write(")")
				if s != " ":
					count_icall_xor_width_under += 1
			'''
			if i in [icall_argPointer,icall_argImm] or i in [icall_argPointer,icall_xor] or i in [icall_argImm,icall_xor]:
				count_both_two += 1

			if i in [icall_argPointer, icall_argImm, icall_xor]:
				count_both_three += 1
			'''

			if i not in compareHandler.icall_argPointer and i not in compareHandler.icall_argImm and \
					i not in compareHandler.icall_xor:
				count_icall_width_under_other += 1

			if i in compareHandler.ind_addr2source:
				out.write(compareHandler.ind_addr2source[i][1])

			out.write("\n")

	count_icall_width_correct_0 = 0
	if len(compareHandler.icall_width_correct) > 0:
		out.write("\ticall argument width correct\n")
		for i in compareHandler.icall_width_correct:
			if i in icall_num_correct:
				if len(compareHandler.icall_width_correct[i]) == 1:
					if compareHandler.icall_width_correct[i][0] == 0:
						count_icall_width_correct_0 += 1
						out.write("\t\t" + str(hex(i)) + "\n")
					else:
						if len(compareHandler.icall_width_correct[i]) == icall_num_correct[i][0]:
							count_icall_width_correct += 1
							out.write("\t\t" + str(hex(i)) + "\n")
				else:
					if len(compareHandler.icall_width_correct[i]) == icall_num_correct[i][0]:
						count_icall_width_correct += 1
						out.write("\t\t" + str(hex(i)) + "\n")


	result = [binary, count_icall_temp, count_icall_paramover, count_icall_taken, count_icall_dcall,
			  count_icall_paramunder, count_icall_paramcorrect, count_icall_paramcorrect_0,
			  count_icall_width_over_other, count_prom, icall_imm_width_over, icall_pointer_width_over,
			  icall_xor_width_over,
			  count_icall_width_under_other, count_icall_width_under, count_icall_pointer_width_under,
			  count_icall_imm_width_under,
			  count_icall_xor_width_under, count_icall_width_correct, count_icall_width_correct_0,
			  compareHandler.total_icall, compareHandler.total_icall_0,count_icall_wrapper_under,count_temp_under]
	return result

def run_ghidra(binary_dir, ghidra_result_path):
	timeout = 600
	for binary in os.listdir(binary_dir):
		binary_path = os.path.join(binary_dir, binary)

		ghidra_result_path_temp = os.path.join(ghidra_result_path, binary)
		print(ghidra_result_path+': ' + binary)
		#print("processing " + binary)

		if os.path.isdir(ghidra_result_path):
			
			ghidra_result_file = os.path.join(ghidra_result_path_temp, binary+'-ghidra.txt')
			#if os.path.isfile(ghidra_result_file):
				#continue
			# cmd = "~/installed-app/ghidra_10.0.1_PUBLIC_20210708/ghidra_10.0.1_PUBLIC/support/analyzeHeadless ~/Decompilation/ghidra-project/ yarpgen -scriptpath /home/yanlin/Decompilation/python-script/ -prescript prescript.py -postScript variable.py -overwrite -import " + binary_path
			cmd = "~/software/ghidra_10.0.1_PUBLIC_20210708/ghidra_10.0.1_PUBLIC/support/analyzeHeadless ~/ghidra-project/ yarpgen -scriptpath /home/yanlin/my-projects/obfuscation-arg/python-script/ -postScript get_signature_from_ghidra.py -overwrite -import " + binary_path
			#os.system(cmd)
			ret_code, output, err_output, is_time_expired, elapsed_time = run_cmd(cmd,
																					   time_out=timeout,
																					   num=-1,
																					   memory_limit=None)

			os.system("rm -r /home/yanlin/ghidra-project/*")

			cmd = ' mv ./ghidra-analysis-out/signature/'+binary+'.txt' + ' ' + ghidra_result_file
			os.system(cmd)

def main():
	config_info = get_config()

	source_dir = config_info['source_folder']

	opt = config_info['opt_level']
	compiler = config_info['compiler']
	get_signature = config_info['get_signature']

	groundtruth_folder = config_info['groundtruth_folder']

	only_compare = config_info['only_compare']
	book_name = config_info['book_name']

	for opt in opts:
		if only_compare == 'true':
			if compiler == 'gcc':
					groundtruth_dir = os.path.join(groundtruth_folder, 'clang')
					groundtruth_dir = os.path.join(groundtruth_dir, 'O0')

					

			if compiler == 'clang':
				groundtruth_dir = os.path.join(groundtruth_folder, compiler)
				groundtruth_dir = os.path.join(groundtruth_dir, opt)


			source_folder = os.path.join(source_dir, compiler)
			source_folder = os.path.join(source_folder, opt)

			print(groundtruth_dir)

			os.system('mkdir -p  result/' + compiler  + '/' + opt)


			for binary in os.listdir(source_folder):
				
				binary_path = os.path.join(source_folder, binary)
				llvm_path = binary + "-llvm-info.txt"

				groundtruth_path = os.path.join(groundtruth_dir, llvm_path)
				


				outfile = os.path.join('./result/' + compiler  + '/' + opt,
									   binary)

				print("*************",outfile)

				out = open(outfile, "w")

				compareHandler = Compare()
				compareHandler.read_signature_file(binary_path, groundtruth_path, compiler, opt,
												   'Original',ghidra='false')

				""" FOR SIGNATURE AT THE CALLEE SITE """
				# for original binaries
				func_num_underestimation = compareHandler.get_func_num_underestimation()
				func_num_correct = compareHandler.get_func_num_correct()
				func_num_overestimation = compareHandler.get_func_num_overestimation()
				func_addr2name = compareHandler.get_funcaddr_name()
				#print(func_num_overestimation)
				#print(func_num_underestimation)
				#print(func_num_correct)

				""" FOR SIGNATURE AT THE CALLER SITE """
				icall_num_underestimation = compareHandler.get_icall_num_underestimation()
				icall_num_correct = compareHandler.get_icall_num_correct()
				icall_num_overestimation = compareHandler.get_icall_num_overestimation()

				out.write(binary + "\n")
				book_name1 = book_name  + '_' + opt + '_callee.xlsx'
				print(book_name1)
				exists = os.path.isfile(book_name1)
				print(exists)
				if exists:
					book = openpyxl.load_workbook(book_name1)
				else:
					book = openpyxl.Workbook()
				sheet_name = 'bin'
				if sheet_name not in book.sheetnames:
					book.create_sheet(sheet_name)
					sheet = book.get_sheet_by_name(sheet_name)

					str1 = ['Application', 'count_variadic_under, ','var2nor_over', 'count_varaidc_to_normal', 'count_normal_to_variadic',
							'count_callee_paramover', 'count_callee_paramunder',
							'count_callee_paramcorrect', 'count_callee_paramcorrect_0', 'count_varadic_paramOver', 'count_readRdx',
							'count_rdx_under', 'count_push_under',
							'callee_arg_width_over_48', 'callee_arg_width_prop', 'count_lea',
							'count_paraWidth_overestimization_other', 'callee_arg_width_over',
							'callee_arg_width_under', 'count_arg_width_correct', 'count_arg_width_correct_0',
							'count_paraWidth_underestimization_not_use',
							'count_paraWidth_underestimization_other', '#total_func', '#total_func_0', '#total_variadic']

					for row in range(1):
						for col in range(26):
							# print str1[col]
							sheet.cell(column=col + 1, row=row + 1, value=str1[col])

				else:
					sheet = book.get_sheet_by_name(sheet_name)

				result1 = print_func_comparison_result(func_num_underestimation, func_num_correct,
													   func_num_overestimation, func_addr2name, out,
													   compareHandler, binary)
				if result1[-3] < 10:
					continue
				print(result1)
				for row in range(sheet.max_row, sheet.max_row + 1):
					for col in range(26):
						print(result1[col])
						sheet.cell(column=col + 1, row=row + 1, value=result1[col])

				book.save(book_name1)

				result2 = print_icall_comparison_result(icall_num_underestimation, icall_num_correct,
														icall_num_overestimation, out, compareHandler, binary)



				book_name2 = book_name  + '_' + opt + '_caller.xlsx'
				exists = os.path.isfile(book_name2)
				if exists:
					book = openpyxl.load_workbook(book_name2)
				else:
					book = openpyxl.Workbook()
				sheet_name = 'bin'
				if sheet_name not in book.sheetnames:
					book.create_sheet(sheet_name)
					sheet = book.get_sheet_by_name(sheet_name)

					str1 = ['Application', 'count_icall_temp', 'count_icall_paramover', 'count_icall_taken',
							'count_icall_dcall',
							'count_icall_paramunder', 'count_icall_paramcorrect', 'count_icall_paramcorrect_0',
							'count_icall_width_over_other', 'count_prom',
							'icall_imm_width_over', 'icall_pointer_width_over', 'icall_xor_width_over',
							'count_icall_width_under_other', 'count_icall_width_under',
							'count_icall_pointer_width_under', 'count_icall_imm_width_under',
							'count_icall_xor_width_under', 'count_icall_width_correct', 'count_icall_width_correct_0',
							'#total_icall','#total_icall_0','count_icall_wrapper_under','count_temp_under']

					for row in range(1):
						for col in range(24):
							# print str1[col]
							sheet.cell(column=col + 1, row=row + 1, value=str1[col])

				else:
					sheet = book.get_sheet_by_name(sheet_name)

				for row in range(sheet.max_row, sheet.max_row + 1):
					for col in range(24):
						# print str[col]
						sheet.cell(column=col + 1, row=row + 1, value=result2[col])

				book.save(book_name2)




		# GET THE FUNCTION SIGNATURE RECOVERED FROM TYPEARMOR FOR NON-OBFUSCATED AND OBFUSCATED BINARIES
		if get_signature == 'true':
			if compiler == 'gcc':
				groundtruth_dir = os.path.join(groundtruth_folder, 'clang')
				groundtruth_dir = os.path.join(groundtruth_dir, 'O0')


			if compiler == 'clang':
				groundtruth_dir = os.path.join(groundtruth_folder, compiler)
				groundtruth_dir = os.path.join(groundtruth_dir, opt)



			source_folder = os.path.join(source_dir, compiler)
			source_folder = os.path.join(source_folder, opt)


			for binary in os.listdir(source_folder):

				print(binary)

				llvm_path = binary + "-llvm-info.txt"
				groundtruth_path = os.path.join(groundtruth_dir, llvm_path)
				non_obfuscated_path = os.path.join(source_folder, binary)

				compareHandler = Compare()
				# p1 = Process(target=inc_forever, name='Process_inc_forever')
				p1 = Process(target=compareHandler.obtain_signature,
							 args=(non_obfuscated_path, groundtruth_path, compiler, opt, 'Original',))
				p1.start()

				p1.join(timeout=60)
				p1.terminate()
				os.system('sudo killall -9 ' + binary)
				os.system('sudo rm -f /var/log/syslog')

					

				



				

				
if __name__ == '__main__':
	main()